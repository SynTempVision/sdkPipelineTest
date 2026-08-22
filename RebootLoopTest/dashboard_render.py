#!/usr/bin/env python3
"""
dashboard_render.py - builds the live status dashboard (the 2x2 overview
grid plus each unit's detail page) that reboot_loop_test.py serves via WAMP.

Kept in its own module, separate from the long-running SSH-polling loop,
specifically so it can be hot-reloaded every cycle (see run_cycle()'s
importlib.reload(dashboard_render) call in reboot_loop_test.py) - editing
this file's HTML/CSS/layout takes effect on the very next cycle of whichever
unit's process runs next, without restarting the multi-day test itself.
generate_run_report() in reboot_loop_test.py (the final, once-per-run static
report) also imports several helpers from here, but doesn't need hot-reload
since it only ever runs once, right as that process is about to exit.
"""

import json
import re
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

WWW_ROOT = Path(r"C:\wamp64\www\sdkPipelineTest")
STATUS_DIR = WWW_ROOT / "status"
DASHBOARD_PATH = WWW_ROOT / "dashboard.html"


def _sanitize_for_path(s: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]", "_", s)


def _esc(s):
    return str(s if s is not None else "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _fmt_snapshot_ts(ts_str):
    """Snapshot JSON stores "timestamp" as %Y-%m-%d %H:%M:%S (sortable,
    unambiguous for programmatic use); this reformats it for display as
    requested - MMMDDYY (e.g. Aug2126) plus time, so "Last Update" reads
    clearly on the multi-day detail page without a stray "2026-08-21" ISO
    date. Falls back to the raw string if it doesn't parse as expected."""
    try:
        return datetime.strptime(ts_str, "%Y-%m-%d %H:%M:%S").strftime("%b%d%y %H:%M:%S")
    except (TypeError, ValueError):
        return ts_str or ""


def _fmt_dt(t):
    """Display-only date format (month/day/year) - the master log itself
    keeps storing/parsing timestamps as %Y-%m-%d for sortability and so
    datetime.strptime in generate_run_report keeps working on old rows;
    this is purely for what a reader sees in the report."""
    return t.strftime("%m/%d/%Y %H:%M:%S")


def _x_ticks(t_min, t_span_sec, pad_l, plot_w, plot_h, y_top, n=5):
    """Shared evenly-spaced wall-clock tick marks/labels for the X axis -
    used by both chart types so a reader can line up a reachability drop
    with the same moment on a metric chart below it. First/last ticks are
    explicitly labeled "test start"/"test end" (in addition to the actual
    time) so it's unambiguous these are the run's boundaries, not just
    another sample point."""
    out = ""
    for i in range(n):
        frac = i / (n - 1)
        t = t_min + timedelta(seconds=t_span_sec * frac)
        x = pad_l + frac * plot_w
        anchor = "start" if i == 0 else ("end" if i == n - 1 else "middle")
        label = t.strftime("%m/%d %H:%M:%S")
        if i == 0:
            label = f"test start {label}"
        elif i == n - 1:
            label = f"test end {label}"
        out += f'<line x1="{x:.1f}" y1="{y_top}" x2="{x:.1f}" y2="{y_top + plot_h}" stroke="#f0f0f0" stroke-width="1"/>\n'
        out += f'<text x="{x:.1f}" y="{y_top + plot_h + 16}" font-size="10" text-anchor="{anchor}" fill="#666">{label}</text>\n'
    return out


def _svg_line_chart(title, timestamps, values, color, width=760, height=170, zero_floor=False):
    """Minimal dependency-free SVG line chart (wall-clock time on X). No
    JS/CDN - has to open standalone on a Windows machine with no guaranteed
    internet. Missing values (unreachable cycles) break the line into
    separate polyline segments rather than being plotted as 0, which would
    misrepresent a gap in data as a real reading. X axis is actual elapsed
    time, not cycle index, since cycle length can vary slightly cycle-to-
    cycle (padding drift) and a reader wants to know "when," not "which
    numbered attempt."

    zero_floor=True pins the y-axis bottom to 0 instead of auto-scaling to
    the data's own min - for a mostly-zero event-count metric (USB errors/
    resets, eth0 flaps), auto-scaling produces a nonsensical axis (e.g. all
    values 0 -> padded to -1..1, a metric that can never go negative) and
    makes an occasional spike look like gradual wandering instead of the
    blip-from-baseline it actually is. Continuous gauge metrics (CPU temp,
    CPU%) keep auto-scaling - a 0-70C axis would flatten the real variation
    in a temperature that only ever ranges ~50-70C into a useless sliver."""
    pad_l, pad_r, pad_t, pad_b = 50, 20, 26, 34
    plot_w = width - pad_l - pad_r
    plot_h = height - pad_t - pad_b

    known = [v for v in values if v is not None]
    if not known or not timestamps:
        return f'<div class="chart"><h3>{title}</h3><p class="empty">No data</p></div>'

    t_min, t_max = min(timestamps), max(timestamps)
    if zero_floor:
        y_min, y_max = 0, max(known)
        if y_max <= 0:
            y_max = 1
    else:
        y_min, y_max = min(known), max(known)
        if y_min == y_max:
            y_min -= 1
            y_max += 1
    t_span = (t_max - t_min).total_seconds() or 1
    y_span = (y_max - y_min) or 1

    def sx(t):
        return pad_l + (t - t_min).total_seconds() / t_span * plot_w

    def sy(v):
        return pad_t + plot_h - (v - y_min) / y_span * plot_h

    segments, current = [], []
    for t, v in zip(timestamps, values):
        if v is None:
            if current:
                segments.append(current)
                current = []
            continue
        current.append((t, v))
    if current:
        segments.append(current)

    polylines = "\n".join(
        '<polyline points="{}" fill="none" stroke="{}" stroke-width="2"/>'.format(
            " ".join(f"{sx(t):.1f},{sy(v):.1f}" for t, v in seg), color
        )
        for seg in segments
    )

    # Invisible per-point hit-circles carrying a native SVG <title> (browser
    # tooltip on hover, no JS/CDN needed - keeps this dependency-free per the
    # module docstring). No visible dot - just the line - the hit-circle
    # only needs to be there for hovering, not drawn.
    points = "\n".join(
        f'<circle cx="{sx(t):.1f}" cy="{sy(v):.1f}" r="6" fill="transparent">'
        f'<title>{_fmt_dt(t)}: {v:g}</title></circle>'
        for seg in segments for t, v in seg
    )

    gridlines = ""
    for frac in (0, 0.5, 1):
        y = pad_t + plot_h * (1 - frac)
        val = y_min + y_span * frac
        gridlines += (
            f'<line x1="{pad_l}" y1="{y:.1f}" x2="{width - pad_r}" y2="{y:.1f}" '
            f'stroke="#e0e0e0" stroke-width="1"/>\n'
            f'<text x="{pad_l - 6}" y="{y + 4:.1f}" font-size="10" text-anchor="end" fill="#666">{val:.1f}</text>\n'
        )

    xticks = _x_ticks(t_min, t_span, pad_l, plot_w, plot_h, pad_t)

    return f'''<div class="chart">
  <h3>{title}</h3>
  <svg viewBox="0 0 {width} {height}" width="100%" height="{height}">
    {gridlines}
    {xticks}
    {polylines}
    {points}
  </svg>
</div>'''


def _svg_multi_line_chart(title, timestamps, series, width=760, height=260):
    """Same dependency-free approach as _svg_line_chart, but plots several
    stats on one shared chart instead of one chart per stat.

    series: list of (label, color, raw_values, zero_floor) tuples, each
    raw_values aligned 1:1 with timestamps. Stats this different in scale -
    CPU % (0-100), CPU Temp in C (~50-70), event counts (0-a few), uptime
    in hours (climbs into the hundreds) - can't share a literal y-axis, so
    each series is independently min-max normalized to 0-100% of its own
    range before plotting. That axis has no real unit of its own; every
    plotted point's native SVG <title> tooltip carries the real,
    unnormalized value instead, so hovering still tells you the actual
    number. A series with no real data at all for this unit/run (e.g.
    Self-Recovered Errors is always "n/a" on an old-pipeline unit) is
    silently skipped - not drawn, not legended - rather than showing a dead
    flat line."""
    pad_l, pad_r, pad_t, pad_b = 40, 20, 14, 34
    plot_w = width - pad_l - pad_r
    plot_h = height - pad_t - pad_b

    if not timestamps:
        return f'<div class="chart"><h3>{title}</h3><p class="empty">No data</p></div>'

    t_min, t_max = min(timestamps), max(timestamps)
    t_span = (t_max - t_min).total_seconds() or 1

    def sx(t):
        return pad_l + (t - t_min).total_seconds() / t_span * plot_w

    def sy(norm_v):
        return pad_t + plot_h - norm_v / 100 * plot_h

    legend_items, polylines, points = "", "", ""
    for label, color, raw, zero_floor in series:
        known = [v for v in raw if v is not None]
        if not known:
            continue
        if zero_floor:
            lo, hi = 0, max(known)
            if hi <= 0:
                hi = 1
        else:
            lo, hi = min(known), max(known)
            if lo == hi:
                lo -= 1
                hi += 1
        span = hi - lo

        segments, current = [], []
        for t, v in zip(timestamps, raw):
            if v is None:
                if current:
                    segments.append(current)
                    current = []
                continue
            current.append((t, v, (v - lo) / span * 100))
        if current:
            segments.append(current)

        polylines += "\n".join(
            '<polyline points="{}" fill="none" stroke="{}" stroke-width="2"/>'.format(
                " ".join(f"{sx(t):.1f},{sy(n):.1f}" for t, v, n in seg), color
            )
            for seg in segments
        ) + "\n"
        points += "\n".join(
            f'<circle cx="{sx(t):.1f}" cy="{sy(n):.1f}" r="5" fill="transparent">'
            f'<title>{label}: {v:g} @ {_fmt_dt(t)}</title></circle>'
            for seg in segments for t, v, n in seg
        ) + "\n"
        legend_items += f'<span class="legend-item"><span class="swatch" style="background:{color}"></span>{label}</span>'

    if not legend_items:
        return f'<div class="chart"><h3>{title}</h3><p class="empty">No data</p></div>'

    gridlines = ""
    for frac, label in ((0, "0%"), (0.5, "50%"), (1, "100%")):
        y = pad_t + plot_h * (1 - frac)
        gridlines += (
            f'<line x1="{pad_l}" y1="{y:.1f}" x2="{width - pad_r}" y2="{y:.1f}" '
            f'stroke="#e0e0e0" stroke-width="1"/>\n'
            f'<text x="{pad_l - 6}" y="{y + 4:.1f}" font-size="10" text-anchor="end" fill="#666">{label}</text>\n'
        )

    xticks = _x_ticks(t_min, t_span, pad_l, plot_w, plot_h, pad_t)

    return f'''<div class="chart">
  <h3>{title}</h3>
  <div class="legend">{legend_items}</div>
  <svg viewBox="0 0 {width} {height}" width="100%" height="{height}">
    {gridlines}
    {xticks}
    {polylines}
    {points}
  </svg>
</div>'''


_STATE_COLORS = {"ok": "#4caf50", "fail": "#e53935", "blip": "#bdbdbd"}


def _reachability_strip(timestamps, states, title="Reachable at checkpoint", width=760, height=44):
    """states: one of "ok"/"fail"/"blip" per cycle. "blip" (gray) is for a
    cycle with no real data - e.g. the checkpoint SSH itself timed out, so
    camera-check-reboot-test.sh never ran and every field came back blank.
    That's a missed checkpoint, not evidence the thing being charted actually
    failed - conflating the two made a single SSH timeout look identical to a
    real Seek enumeration failure (see PT1 cycle 221, 2026-08-21 - blank
    Seek Camera Enumeration counted as FAILED when the row was simply
    unreachable that cycle)."""
    n = max(len(timestamps), 1)
    seg_w = width / n
    rects = "\n".join(
        f'<rect x="{i * seg_w:.1f}" y="0" width="{seg_w + 0.5:.1f}" height="24" '
        f'fill="{_STATE_COLORS[s]}"><title>{_fmt_dt(timestamps[i])}: {s}</title></rect>'
        for i, s in enumerate(states)
    )
    t_min, t_max = min(timestamps), max(timestamps)
    t_span = (t_max - t_min).total_seconds() or 1
    xticks = _x_ticks(t_min, t_span, 0, width, 0, 24, n=5)
    return f'''<div class="chart">
  <h3>{title}</h3>
  <svg viewBox="0 0 {width} {height}" width="100%" height="{height}">{rects}{xticks}</svg>
</div>'''


_TRENDED_METRICS = [
    # (column name, color, zero_floor) - zero_floor for event counts that
    # can't go negative (an occasional spike should read as a spike off a
    # true zero baseline, not as "wandering" within an auto-scaled range).
    ("Total CPU %", "#00838f", False),
    ("Pipeline CPU %", "#558b2f", False),
    ("CPU Temp (C)", "#e65100", False),
    ("Pipeline Restarts (24h)", "#ef6c00", True),
    ("Self-Recovered Errors (24h)", "#3949ab", True),
    ("USB Communication Errors", "#1565c0", True),
    ("USB Resets", "#6a1b9a", True),
    ("Ethernet Flap Count", "#2e7d32", True),
]


def _unit_trend_htmls(log_path, run_id, combined_width=760, combined_height=260):
    """Reads a unit's own log file back, filtered to the current run, and
    builds the live trend views generate_run_report() produces for a
    finished run - but on whatever rows exist so far. Returns
    (strips_html, combined_html, separate_html), each "" if the log isn't
    there yet or has fewer than 2 rows for this run (a single point can't
    show a trend, and the very first cycle of a fresh run hits this every
    time).

    strips_html: the categorical reachable/Seek/image.pgm pass-fail strips
    - only meaningful on the per-unit detail page, not compact enough for
    the 4-way overview grid.

    combined_html: every trended numeric stat (see _TRENDED_METRICS) on one
    shared chart, independently normalized so unlike scales (CPU %, CPU
    Temp in C, event counts) can share an axis - compact enough for the
    overview grid's 4-up layout.

    separate_html: the same stats, but one real-units chart per stat - used
    on the per-unit detail page, which has room for it and doesn't need the
    normalization tradeoff (real y-axis values, no "hover to see the actual
    number" indirection)."""
    try:
        wb = openpyxl.load_workbook(log_path, read_only=True, data_only=True)
        ws = wb.active
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        run_id_idx = header.index("Run ID")
        rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[run_id_idx] == run_id]
        wb.close()
    except Exception:
        # Broad on purpose: this reads another unit's log file, which that
        # unit's own process can be mid-write on at this exact moment (its
        # log_row() truncates and rewrites the whole .xlsx via wb.save()) -
        # caught this in practice as zipfile.BadZipFile ("File is not a zip
        # file") on 2026-08-21, which crashed the whole reading process since
        # BadZipFile isn't an OSError/KeyError/ValueError/StopIteration. This
        # chart is always best-effort - any read failure here should just
        # skip this cycle's trend, never take down the test itself.
        return "", "", ""
    if len(rows) < 2:
        return "", "", ""

    timestamp_idx = header.index("Timestamp")
    reachable_col = header.index("Reachable at checkpoint")
    timestamps = [datetime.strptime(r[timestamp_idx], "%Y-%m-%d %H:%M:%S") for r in rows]
    reachable_flags = [r[reachable_col] == "Yes" for r in rows]

    def col(col_name):
        return [r[header.index(col_name)] for r in rows]

    def numeric_series(col_name):
        if col_name not in header:
            return [None] * len(rows)  # older log predating this column - treat as no data, not a crash
        idx = header.index(col_name)
        out = []
        for r in rows:
            try:
                out.append(float(r[idx]))
            except (TypeError, ValueError):
                out.append(None)
        return out

    def make_states(raw_values, reachable, is_ok):
        states = []
        for v, r in zip(raw_values, reachable):
            if not r or v in (None, ""):
                states.append("blip")
            else:
                states.append("ok" if is_ok(v) else "fail")
        return states

    seek_states = make_states(col("Seek Camera Enumeration"), reachable_flags, lambda v: v == "OK")
    image_states = make_states(col("image.pgm Healthy"), reachable_flags, lambda v: v == "Yes")

    strips_html = _reachability_strip(timestamps, ["ok" if r else "fail" for r in reachable_flags], "Reachable at checkpoint")
    strips_html += "\n" + _reachability_strip(timestamps, seek_states, "Seek Camera Enumeration (green=OK, red=FAILED, gray=no data)")
    strips_html += "\n" + _reachability_strip(timestamps, image_states, "image.pgm Healthy (green=Yes, red=No, gray=no data)")

    series = [(name, color, numeric_series(name), zero_floor) for name, color, zero_floor in _TRENDED_METRICS]
    combined_html = _svg_multi_line_chart(
        "All stats (each line normalized 0-100% of its own range - hover a point for the real value)",
        timestamps, series, width=combined_width, height=combined_height,
    )

    separate_html = "\n".join(
        _svg_line_chart(name, timestamps, numeric_series(name), color, zero_floor=zero_floor)
        for name, color, zero_floor in _TRENDED_METRICS
    )

    return strips_html, combined_html, separate_html


_DASHBOARD_CSS = '''
  body { font-family: Arial, sans-serif; margin: 24px; color: #222; }
  h1 { font-size: 18px; }
  h2 { font-size: 15px; border-top: 2px solid #eee; padding-top: 18px; margin-top: 28px; }
  a { color: #1565c0; }
  .updated { color: #777; font-size: 12px; margin-bottom: 16px; }
  .back { display: inline-block; margin-bottom: 14px; font-size: 13px; }
  table { border-collapse: collapse; width: 100%; font-size: 13px; }
  th, td { padding: 6px 10px; text-align: left; border-bottom: 1px solid #eee; white-space: nowrap; }
  th { background: #f5f5f5; position: sticky; top: 0; }
  td.unit { font-weight: bold; }
  .ok { color: #2e7d32; }
  .bad { color: #c62828; font-weight: bold; }
  .empty { color: #999; }
  .chart { margin-bottom: 22px; max-width: 760px; }
  .chart h3 { margin: 0 0 6px 0; font-size: 13px; color: #444; }
  .legend { font-size: 11px; margin-bottom: 4px; }
  .legend-item { display: inline-flex; align-items: center; margin-right: 12px; color: #444; }
  .swatch { display: inline-block; width: 10px; height: 10px; border-radius: 2px; margin-right: 4px; }
  dl.summary { display: grid; grid-template-columns: max-content 1fr; gap: 4px 16px; font-size: 13px; max-width: 640px; }
  dl.summary dt { color: #666; }
  dl.summary dd { margin: 0; }
  .grid { display: grid; grid-template-columns: 1fr 1fr; gap: 14px; }
  .cell { border: 1px solid #e0e0e0; border-radius: 6px; padding: 12px 14px; min-width: 0; }
  .cell h2 { margin: 0 0 6px 0; padding: 0; border: none; font-size: 15px; }
  .cell h2 a { text-decoration: none; }
  .cell .chart { margin-bottom: 6px; max-width: none; }
  .cell-stats { font-size: 12px; color: #444; margin-bottom: 8px; display: flex; flex-wrap: wrap; gap: 4px 14px; }
  .cell-stats b { color: #222; }
  @media (max-width: 1000px) { .grid { grid-template-columns: 1fr; } }
'''


def _unit_detail_filename(unit_label):
    return f"unit_{_sanitize_for_path(unit_label)}.html"


def render_unit_detail(snapshot):
    """One standalone page per unit - the drill-down target when you click a
    unit on the overview grid. Each process only ever owns its own unit's
    snapshot, so it only ever writes its own detail page here; no cross-
    process file contention, same as write_status_snapshot()."""
    s = snapshot
    reachable_ok = s.get("reachable") is True
    status_text = s.get("status") or ("" if reachable_ok else "UNREACHABLE")
    status_class = "ok" if (reachable_ok and (not status_text or status_text == "PIPELINE HEALTHY")) else "bad"
    img_class = "ok" if s.get("image_healthy") else "bad"

    strips, separate = "", ""
    if s.get("log_path") and s.get("run_id"):
        strips, _, separate = _unit_trend_htmls(s["log_path"], s["run_id"])
    charts = (strips + "\n" + separate) if (strips or separate) else '<p class="empty">Not enough cycles yet for a trend (need at least 2).</p>'

    html = f'''<!doctype html>
<html>
<head>
<meta charset="utf-8">
<title>{_esc(s.get("unit"))} - Test 0 Pipeline Monitoring</title>
<style>{_DASHBOARD_CSS}</style>
</head>
<body>
<a class="back" href="dashboard.html">&larr; back to all units</a>
<h1>{_esc(s.get("unit"))}</h1>
<div class="updated">Rewritten fresh every cycle - reload this page (F5) to see the latest. Generated {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.</div>
<dl class="summary">
  <dt>IP</dt><dd>{_esc(s.get("ip"))}</dd>
  <dt>Pipeline</dt><dd>{_esc(s.get("pipeline"))}</dd>
  <dt>Action</dt><dd>{_esc(s.get("action"))}</dd>
  <dt>Cycle</dt><dd>{_esc(s.get("cycle"))}</dd>
  <dt>Last Update</dt><dd>{_esc(_fmt_snapshot_ts(s.get("timestamp")))}</dd>
  <dt>Reachable</dt><dd class="{'ok' if reachable_ok else 'bad'}">{"Yes" if reachable_ok else "No"}</dd>
  <dt>Seek Enum</dt><dd class="{'ok' if s.get('seek_enumeration') == 'OK' else 'bad'}">{_esc(s.get("seek_enumeration"))}</dd>
  <dt>image.pgm</dt><dd class="{img_class}">{_esc(s.get("image_desc"))}</dd>
  <dt>Restarts (24h)</dt><dd>{_esc(s.get("pipeline_restarts"))}</dd>
  <dt>Self-Recovered</dt><dd>{_esc(s.get("self_recovered"))}</dd>
  <dt>Wrapper Alive</dt><dd>{_esc(s.get("wrapper_alive"))}</dd>
  <dt>CPU %</dt><dd>{_esc(s.get("cpu_pct"))}%</dd>
  <dt>CPU Temp</dt><dd>{_esc(s.get("cpu_temp"))}&deg;C</dd>
  <dt>Uptime</dt><dd>{_esc(s.get("uptime_hours"))}h</dd>
  <dt>Throttled</dt><dd>{_esc(s.get("throttled_meaning"))}</dd>
  <dt>Status</dt><dd class="{status_class}">{_esc(status_text)}</dd>
</dl>
<h2>Trends</h2>
{charts}
</body>
</html>
'''
    (WWW_ROOT / _unit_detail_filename(s.get("unit"))).write_text(html, encoding="utf-8")


def _unit_sort_key(unit_label):
    """Pairs each old/new unit onto one row - new-pipeline unit on the left,
    its old-pipeline counterpart on the right (e.g. PT0.1NEW, PT0.1, then
    PT0.2NEW, PT0.2) - so the two units under direct comparison sit side by
    side instead of scattered across the grid by alphabetical filename."""
    m = re.match(r"PT(\d+(?:\.\d+)?)(NEW)?", unit_label)
    if not m:
        return (999, 0, unit_label)
    return (float(m.group(1)), 0 if m.group(2) else 1, unit_label)


def render_dashboard():
    """One-screen 2x2 grid, one cell per unit - a compact stat line plus
    that unit's combined trend graph, sized down (see _unit_trend_htmls'
    combined_width/height) so all 4 fit without scrolling on a normal
    monitor. Full detail (the categorical strips, every raw field) lives
    one click away on each unit's own page - see render_unit_detail()."""
    snapshots = []
    if STATUS_DIR.exists():
        for f in sorted(STATUS_DIR.glob("*.json")):
            try:
                snapshots.append(json.loads(f.read_text(encoding="utf-8")))
            except (json.JSONDecodeError, OSError):
                continue  # skip a snapshot caught mid-write rather than crash the whole dashboard
    snapshots.sort(key=lambda s: _unit_sort_key(s.get("unit", "")))

    cells_html = ""
    for s in snapshots:
        reachable_ok = s.get("reachable") is True
        status_text = s.get("status") or ("" if reachable_ok else "UNREACHABLE")
        status_class = "ok" if (reachable_ok and (not status_text or status_text == "PIPELINE HEALTHY")) else "bad"
        img_class = "ok" if s.get("image_healthy") else "bad"
        detail_href = _unit_detail_filename(s.get("unit"))

        combined = ""
        if s.get("log_path") and s.get("run_id"):
            _, combined, _ = _unit_trend_htmls(s["log_path"], s["run_id"], combined_width=560, combined_height=190)
        if not combined:
            combined = '<p class="empty">Not enough cycles yet for a trend (need at least 2).</p>'

        cells_html += f'''<div class="cell">
  <h2><a href="{detail_href}">{_esc(s.get("unit"))}</a></h2>
  <div class="cell-stats">
    <span><b>{_esc(s.get("pipeline"))}</b> pipeline</span>
    <span>Cycle <b>{_esc(s.get("cycle"))}</b></span>
    <span>Reachable <b class="{'ok' if reachable_ok else 'bad'}">{"Yes" if reachable_ok else "No"}</b></span>
    <span>Seek <b class="{'ok' if s.get('seek_enumeration') == 'OK' else 'bad'}">{_esc(s.get("seek_enumeration"))}</b></span>
    <span>image.pgm <b class="{img_class}">{"OK" if s.get("image_healthy") else "BAD"}</b></span>
    <span>CPU <b>{_esc(s.get("cpu_pct"))}%</b> / <b>{_esc(s.get("cpu_temp"))}&deg;C</b></span>
    <span>Uptime <b>{_esc(s.get("uptime_hours"))}h</b></span>
    <span class="{status_class}">{_esc(status_text)}</span>
  </div>
  {combined}
</div>'''

    html = f'''<!doctype html>
<html>
<head>
<meta charset="utf-8">
<title>Test 0 Pipeline Monitoring - Live Status</title>
<style>{_DASHBOARD_CSS}</style>
</head>
<body>
<h1>Test 0 Pipeline Monitoring - Live Status</h1>
<div class="updated">Rewritten fresh every cycle by whichever unit finishes next - reload this page (F5) to see the latest. Click a unit for its full detail page. Generated {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.</div>
{'<p class="empty">No status snapshots yet - waiting for the first cycle to complete.</p>' if not snapshots else f'<div class="grid">{cells_html}</div>'}
</body>
</html>
'''
    DASHBOARD_PATH.write_text(html, encoding="utf-8")
