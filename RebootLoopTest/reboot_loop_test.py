#!/usr/bin/env python3
"""
reboot_loop_test.py - drives a fixed-cadence reboot OR pipeline-restart loop
against a single camera and logs a snapshot of its state each cycle. Built for
comparing a plain PT unit against a PT unit running the RAM OS overlay.

Two modes, selected with --action:

  reboot (Test 1, 2-minute cycle):
    issue `sudo reboot` (fire-and-forget - the connection dies with the reboot),
    sleep to T+1:30, SSH back in (capped at 20s) and check, log, pad to T+2:00,
    repeat immediately.

  restart (Test 2, 60-second cycle):
    run `pipeline_restart.sh` as a BLOCKING command (observed to take ~30s) -
    the wait for it to return already covers "give it time to recover", so we
    check immediately after rather than sleeping separately for that purpose.
    Then pad whatever's left to reach 60s total, purely for uniform cadence
    (not for recovery), log, repeat immediately.

Common to both: if the checkpoint SSH attempt fails/times out, log it as
unreachable and continue the loop anyway - a stuck/dead cycle must never stop
the test from gathering further data points. Reuses this project's existing
SSH connection pattern (see run_checks_v2.py) rather than reinventing it.

Requires: pip install paramiko openpyxl

Usage:
  python reboot_loop_test.py --ip 172.17.103.5 --user camera --password <pw> --unit "PT" --action reboot --pipeline old
  python reboot_loop_test.py --ip 172.17.103.5 --user camera --password <pw> --unit "PT RamOS" --action restart --pipeline old --cycles 60
  python reboot_loop_test.py --ip 172.17.103.135 --password <pw> --unit "PT" --action restart --pipeline new --duration-min 60
  python reboot_loop_test.py --ip ... --action reboot --pipeline old --duration-min 60   # run for ~1 hour instead of a fixed count
"""

import argparse
import json
import re
import socket
import sys
import time
from collections import Counter
from datetime import datetime, timedelta
from pathlib import Path

import paramiko
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

CHECK_SCRIPT = Path(__file__).parent / "camera-check-reboot-test.sh"
CHECK_TIMEOUT_SEC = 20        # hard cap on the checkpoint SSH check itself
CONNECT_TIMEOUT_SEC = 5

# Test 1 (reboot) timing
REBOOT_CHECK_INTERVAL_SEC = 90   # T+1:30 - when we check in each cycle
REBOOT_CYCLE_LENGTH_SEC = 120    # T+2:00 - total nominal cycle length

# Test 2 (pipeline_restart) timing
RESTART_TIMEOUT_SEC = 45         # cap on the blocking pipeline_restart.sh call (~30s observed)
RESTART_CYCLE_LENGTH_SEC = 60    # total nominal cycle length

# Test 0 (steady-state endurance) timing - no reboot/restart trigger at all,
# just a passive check on a fixed cadence, to catch slow-building problems a
# constantly-restarting stress loop would mask by resetting state every cycle.
NONE_CYCLE_LENGTH_SEC = 300     # Test 0: 5-minute steady-state cycle

COLUMNS = [
    "Run ID", "Cycle", "Timestamp", "Unit", "Action", "Reachable at checkpoint", "Unreachable Reason",
    "Seek Camera Enumeration", "Enum Fail Time (s since boot)",
    "USB Communication Errors", "USB Resets", "Ethernet Flap Count",
    "image.pgm Size", "image.pgm Age (s)", "image.pgm Healthy",
    "Pipeline Type", "Pipeline Version", "Pipeline Restarts (24h)", "Self-Recovered Errors (24h)",
    "Wrapper Script Alive",
    "Uptime (Hours)",
    "Total CPU %", "Pipeline CPU %",
    "Calibration", "CPU Temp (C)", "Throttled", "Throttled (Meaning)",
    "imgserv Port", "RTSP 8554 Port", "Network (gateway ping)",
    "RAM (total,used,free,available,buff/cache)", "Overlay Active",
    "Overlay Usage (size,used,avail,pct)", "Status", "Raw dmesg tail",
]

BASE_FONT = Font(name="Arial")
HEADER_FONT = Font(name="Arial", bold=True)
HEADER_FILL = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
RED_FILL = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
GREEN_FILL = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
THIN = Side(style="thin", color="FF000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def connect(ip, user, password, timeout=CONNECT_TIMEOUT_SEC):
    sock = socket.create_connection((ip, 22), timeout=timeout)
    transport = paramiko.Transport(sock, disabled_algorithms=dict(
        pubkeys=["rsa-sha2-256", "rsa-sha2-512"],
        kex=["diffie-hellman-group-exchange-sha256", "diffie-hellman-group16-sha512"],
    ))
    transport.set_keepalive(10)
    transport.connect(username=user, password=password)
    return transport


def run_remote_command(transport, command, read_timeout=CHECK_TIMEOUT_SEC):
    channel = transport.open_session()
    channel.settimeout(read_timeout)
    channel.exec_command(command)
    channel.shutdown_write()
    output = b""
    try:
        while True:
            chunk = channel.recv(4096)
            if not chunk:
                break
            output += chunk
    except socket.timeout:
        pass
    channel.close()
    return output.decode(errors="replace").strip()


def trigger_reboot(ip, user, password):
    """Fire-and-forget: issue reboot, don't wait for a clean response since the
    connection will be torn down by the reboot itself."""
    try:
        transport = connect(ip, user, password, timeout=CONNECT_TIMEOUT_SEC)
        try:
            channel = transport.open_session()
            channel.exec_command("sudo reboot")
            time.sleep(1)  # give the command a moment to actually fire before we tear down
        finally:
            transport.close()
        return True, None
    except Exception as e:
        return False, f"{type(e).__name__}: {e}"


def trigger_restart_blocking(ip, user, password, pipeline):
    """BLOCKING: restart the pipeline and wait for it to actually finish, so
    the wait itself satisfies "give it time to recover" - no separate
    recovery sleep needed for the old-pipeline path, only cadence padding.
    Capped at RESTART_TIMEOUT_SEC so a stuck command can't hang the test.

    old: pipeline_restart.sh itself blocks for ~30s, covering the recovery
    wait on its own.
    new: `systemctl restart` returns as soon as the unit (re)starts, which
    is much faster than frames actually flowing again - add an explicit
    10s settle sleep after it returns, matching the same restart_pipeline()
    precedent in camera-check-v2.sh. Also runs `reset-failed` right before
    the restart - this test deliberately restarts every 60s, which trips
    systemd's StartLimitBurst (50 restarts/rolling hour) around cycle 51
    and silently no-ops every restart after that (confirmed live on sv-206,
    2026-08-20 - image.pgm stuck MISSING for the rest of the run even
    though `systemctl restart` reports success). reset-failed clears the
    burst counter each cycle so this test measures actual pipeline
    resilience instead of re-discovering the same rate limit every hour."""
    command = (
        "sudo bash /usr/local/bin/camera/pipeline_restart.sh" if pipeline == "old"
        else "sudo systemctl reset-failed seekcamera-gstreamer; sudo systemctl restart seekcamera-gstreamer"
    )
    try:
        transport = connect(ip, user, password, timeout=CONNECT_TIMEOUT_SEC)
        try:
            output = run_remote_command(transport, command, read_timeout=RESTART_TIMEOUT_SEC)
            if pipeline == "new":
                time.sleep(10)
        finally:
            transport.close()
        return True, output
    except Exception as e:
        return False, f"{type(e).__name__}: {e}"


def run_check(ip, user, password):
    """The T+1:30 checkpoint: connect, push the check script, run it, parse output.
    Capped at CHECK_TIMEOUT_SEC total - a stuck cycle must not hang the whole test."""
    transport = connect(ip, user, password, timeout=CONNECT_TIMEOUT_SEC)
    try:
        sftp = paramiko.SFTPClient.from_transport(transport)
        sftp.put(str(CHECK_SCRIPT), "/tmp/camera-check-reboot-test.sh")
        sftp.close()
        raw = run_remote_command(transport, "bash /tmp/camera-check-reboot-test.sh", read_timeout=CHECK_TIMEOUT_SEC)
        fields = parse_output(raw)
        if not fields:
            # SSH/SFTP succeeded but the check script produced no parseable
            # KEY=VALUE output - e.g. the read hit CHECK_TIMEOUT_SEC before
            # the script printed anything, or the remote command failed
            # outright. This must NOT be logged as a normal, data-complete
            # cycle - see PT2 cycle 429, 2026-08-21: it landed in the master
            # log as Reachable=Yes with every other field blank, which
            # fooled the report's "first Seek enumeration failure" detection
            # into picking this row instead of the real failure two cycles
            # later. Raising here routes it through the same reachability-
            # failure path as a dropped connection, so it's logged and
            # treated consistently instead of masquerading as success.
            raise RuntimeError(f"check script produced no parseable output (raw={raw[:200]!r})")
        return fields
    finally:
        transport.close()


def parse_output(raw):
    fields = {}
    for line in raw.splitlines():
        if "=" in line:
            key, _, val = line.partition("=")
            fields[key.strip()] = val.strip()
    return fields


# Bit layout per vcgencmd get_throttled - ported from run_checks_v2.py's
# decode_throttled so both tools read it the same way.
THROTTLE_BITS = [(0x1, "Under-V"), (0x2, "FreqCap"), (0x4, "Throttle"), (0x8, "Hot")]
THROTTLE_HIST_BITS = [(0x10000, "Under-V"), (0x20000, "FreqCap"), (0x40000, "Throttle"), (0x80000, "Hot")]


def decode_throttled(hex_str):
    if not hex_str:
        return ""
    try:
        val = int(hex_str, 16)
    except ValueError:
        return hex_str
    if val == 0:
        return "Clean"
    active = [label for mask, label in THROTTLE_BITS if val & mask]
    hist = [label for mask, label in THROTTLE_HIST_BITS if val & mask]
    parts = []
    if active:
        parts.append("+".join(active))
    if hist:
        parts.append(f"({'+'.join(hist)} hist)")
    return " ".join(parts)


def _uptime_hours(uptime_sec):
    try:
        return round(float(uptime_sec) / 3600, 1)
    except (TypeError, ValueError):
        return ""


def sync_header(ws):
    """Inserts any column newly added to COLUMNS that an existing sheet's
    header row doesn't have yet, at the position COLUMNS says it belongs -
    insert_cols shifts every existing row's cells too, so historical rows
    stay aligned under their original headers instead of drifting once a
    new column lands in the middle of the layout instead of at the end.
    Ported from run_checks_v2.py's identical fix - see
    RebootLoopTest_PT1/PT2's 2026-08-20 column-misalignment repair for why
    this matters: adding Calibration and Run ID to COLUMNS without this
    silently shifted every row appended to an already-existing log file."""
    existing = [c.value for c in ws[1]]
    if existing == COLUMNS:
        return
    for idx, name in enumerate(COLUMNS, start=1):
        if idx > len(existing) or existing[idx - 1] != name:
            ws.insert_cols(idx)
            ws.cell(row=1, column=idx, value=name)
            existing.insert(idx - 1, name)


def ensure_log_file(log_path):
    if log_path.exists():
        wb = openpyxl.load_workbook(log_path)
        ws = wb.active
        sync_header(ws)
        wb.save(log_path)
        return wb, ws
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reboot Loop Test"
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    wb.save(log_path)
    return wb, ws


def log_row(log_path, row_values, reachable):
    wb, ws = ensure_log_file(log_path)
    ws.append(row_values)
    row_idx = ws.max_row
    fill = GREEN_FILL if reachable else RED_FILL
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = BASE_FONT
        cell.border = BORDER
        cell.fill = fill
    wb.save(log_path)


WWW_ROOT = Path(r"C:\wamp64\www\sdkPipelineTest")
STATUS_DIR = WWW_ROOT / "status"
DASHBOARD_PATH = WWW_ROOT / "dashboard.html"


def write_status_snapshot(unit_label, ip, pipeline, action, run_id, cycle_num, reachable, fields, log_path):
    """Each of the (up to) 4 running processes owns exactly one file here -
    no write contention, no locking needed. The dashboard is rebuilt by
    reading whichever snapshot files currently exist, so it stays fresh as
    long as at least one process is still cycling."""
    STATUS_DIR.mkdir(exist_ok=True)
    snapshot = {
        "unit": unit_label,
        "ip": ip,
        "pipeline": pipeline,
        "action": action,
        "run_id": run_id,
        "cycle": cycle_num,
        "log_path": str(log_path),
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "reachable": reachable,
        "seek_enumeration": fields.get("SEEK_ENUMERATION", ""),
        "image_healthy": fields.get("IMAGE_OK") == "1",
        "image_desc": fields.get("IMAGE_PGM", ""),
        "pipeline_type": fields.get("PIPELINE_TYPE", ""),
        "pipeline_restarts": fields.get("PIPELINE_RESTARTS", ""),
        "self_recovered": fields.get("SELF_RECOVERED_ERRORS", ""),
        "wrapper_alive": fields.get("WRAPPER_ALIVE", ""),
        "cpu_pct": fields.get("TOTAL_CPU_PCT", ""),
        "cpu_temp": fields.get("TEMP_C", ""),
        "throttled_meaning": decode_throttled(fields.get("THROTTLED", "")),
        "status": fields.get("STATUS", ""),
    }
    tmp = STATUS_DIR / f"{unit_label}.json.tmp"
    tmp.write_text(json.dumps(snapshot, indent=2), encoding="utf-8")
    tmp.replace(STATUS_DIR / f"{unit_label}.json")  # atomic - dashboard render never sees a half-written file
    return snapshot


def _dashboard_unit_charts(log_path, run_id):
    """Reads a unit's own log file back, filtered to the current run, and
    builds the same reachability strips / line charts generate_run_report()
    produces for a finished run - but live, on whatever rows exist so far.
    Returns "" (caller falls back to a placeholder) if the log isn't there
    yet or has fewer than 2 rows for this run - a single point can't show a
    trend, and the very first cycle of a fresh run hits this every time."""
    try:
        wb = openpyxl.load_workbook(log_path, read_only=True, data_only=True)
        ws = wb.active
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        run_id_idx = header.index("Run ID")
        rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[run_id_idx] == run_id]
        wb.close()
    except (OSError, KeyError, ValueError, StopIteration):
        return ""
    if len(rows) < 2:
        return ""

    timestamp_idx = header.index("Timestamp")
    reachable_col = header.index("Reachable at checkpoint")
    timestamps = [datetime.strptime(r[timestamp_idx], "%Y-%m-%d %H:%M:%S") for r in rows]
    reachable_flags = [r[reachable_col] == "Yes" for r in rows]

    def col(col_name):
        return [r[header.index(col_name)] for r in rows]

    def numeric_series(col_name):
        idx = header.index(col_name)
        out = []
        for r in rows:
            try:
                out.append(float(r[idx]))
            except (TypeError, ValueError):
                out.append(None)
        return out

    def make_states(raw_values, reachable, is_ok):
        states = []
        for v, r in zip(raw_values, reachable):
            if not r or v in (None, ""):
                states.append("blip")
            else:
                states.append("ok" if is_ok(v) else "fail")
        return states

    seek_states = make_states(col("Seek Camera Enumeration"), reachable_flags, lambda v: v == "OK")
    image_states = make_states(col("image.pgm Healthy"), reachable_flags, lambda v: v == "Yes")

    html = _reachability_strip(timestamps, ["ok" if r else "fail" for r in reachable_flags], "Reachable at checkpoint")
    html += "\n" + _reachability_strip(timestamps, seek_states, "Seek Camera Enumeration (green=OK, red=FAILED, gray=no data)")
    html += "\n" + _reachability_strip(timestamps, image_states, "image.pgm Healthy (green=Yes, red=No, gray=no data)")
    for col_name, color in (("CPU Temp (C)", "#e65100"), ("Total CPU %", "#00838f")):
        html += "\n" + _svg_line_chart(col_name, timestamps, numeric_series(col_name), color)
    for col_name, color in (("USB Communication Errors", "#1565c0"), ("USB Resets", "#6a1b9a"), ("Ethernet Flap Count", "#2e7d32")):
        html += "\n" + _svg_line_chart(col_name, timestamps, numeric_series(col_name), color, zero_floor=True)
    return html


_DASHBOARD_CSS = '''
  body { font-family: Arial, sans-serif; margin: 24px; color: #222; }
  h1 { font-size: 18px; }
  h2 { font-size: 15px; border-top: 2px solid #eee; padding-top: 18px; margin-top: 28px; }
  a { color: #1565c0; }
  .updated { color: #777; font-size: 12px; margin-bottom: 16px; }
  .back { display: inline-block; margin-bottom: 14px; font-size: 13px; }
  table { border-collapse: collapse; width: 100%; font-size: 13px; }
  th, td { padding: 6px 10px; text-align: left; border-bottom: 1px solid #eee; white-space: nowrap; }
  th { background: #f5f5f5; position: sticky; top: 0; }
  td.unit { font-weight: bold; }
  .ok { color: #2e7d32; }
  .bad { color: #c62828; font-weight: bold; }
  .empty { color: #999; }
  .chart { margin-bottom: 22px; max-width: 760px; }
  .chart h3 { margin: 0 0 6px 0; font-size: 13px; color: #444; }
  dl.summary { display: grid; grid-template-columns: max-content 1fr; gap: 4px 16px; font-size: 13px; max-width: 640px; }
  dl.summary dt { color: #666; }
  dl.summary dd { margin: 0; }
'''


def _unit_detail_filename(unit_label):
    return f"unit_{_sanitize_for_path(unit_label)}.html"


def render_unit_detail(snapshot):
    """One standalone page per unit - the drill-down target when you click a
    unit on the overview table. Each process only ever owns its own unit's
    snapshot, so it only ever writes its own detail page here; no cross-
    process file contention, same as write_status_snapshot()."""
    s = snapshot
    reachable_ok = s.get("reachable") is True
    status_text = s.get("status") or ("" if reachable_ok else "UNREACHABLE")
    status_class = "ok" if (reachable_ok and (not status_text or status_text == "PIPELINE HEALTHY")) else "bad"
    img_class = "ok" if s.get("image_healthy") else "bad"

    charts = ""
    if s.get("log_path") and s.get("run_id"):
        charts = _dashboard_unit_charts(s["log_path"], s["run_id"])
    if not charts:
        charts = '<p class="empty">Not enough cycles yet for a trend (need at least 2).</p>'

    html = f'''<!doctype html>
<html>
<head>
<meta charset="utf-8">
<title>{_esc(s.get("unit"))} - Pipeline Test</title>
<style>{_DASHBOARD_CSS}</style>
</head>
<body>
<a class="back" href="dashboard.html">&larr; back to all units</a>
<h1>{_esc(s.get("unit"))}</h1>
<div class="updated">Rewritten fresh every cycle - reload this page (F5) to see the latest. Generated {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.</div>
<dl class="summary">
  <dt>IP</dt><dd>{_esc(s.get("ip"))}</dd>
  <dt>Pipeline</dt><dd>{_esc(s.get("pipeline"))}</dd>
  <dt>Action</dt><dd>{_esc(s.get("action"))}</dd>
  <dt>Cycle</dt><dd>{_esc(s.get("cycle"))}</dd>
  <dt>Last Update</dt><dd>{_esc(s.get("timestamp"))}</dd>
  <dt>Reachable</dt><dd class="{'ok' if reachable_ok else 'bad'}">{"Yes" if reachable_ok else "No"}</dd>
  <dt>Seek Enum</dt><dd class="{'ok' if s.get('seek_enumeration') == 'OK' else 'bad'}">{_esc(s.get("seek_enumeration"))}</dd>
  <dt>image.pgm</dt><dd class="{img_class}">{_esc(s.get("image_desc"))}</dd>
  <dt>Restarts (24h)</dt><dd>{_esc(s.get("pipeline_restarts"))}</dd>
  <dt>Self-Recovered</dt><dd>{_esc(s.get("self_recovered"))}</dd>
  <dt>Wrapper Alive</dt><dd>{_esc(s.get("wrapper_alive"))}</dd>
  <dt>CPU %</dt><dd>{_esc(s.get("cpu_pct"))}%</dd>
  <dt>CPU Temp</dt><dd>{_esc(s.get("cpu_temp"))}&deg;C</dd>
  <dt>Throttled</dt><dd>{_esc(s.get("throttled_meaning"))}</dd>
  <dt>Status</dt><dd class="{status_class}">{_esc(status_text)}</dd>
</dl>
<h2>Trends</h2>
{charts}
</body>
</html>
'''
    (WWW_ROOT / _unit_detail_filename(s.get("unit"))).write_text(html, encoding="utf-8")


def render_dashboard():
    snapshots = []
    if STATUS_DIR.exists():
        for f in sorted(STATUS_DIR.glob("*.json")):
            try:
                snapshots.append(json.loads(f.read_text(encoding="utf-8")))
            except (json.JSONDecodeError, OSError):
                continue  # skip a snapshot caught mid-write rather than crash the whole dashboard

    rows_html = ""
    for s in snapshots:
        reachable_ok = s.get("reachable") is True
        status_text = s.get("status") or ("" if reachable_ok else "UNREACHABLE")
        status_class = "ok" if (reachable_ok and (not status_text or status_text == "PIPELINE HEALTHY")) else "bad"
        img_class = "ok" if s.get("image_healthy") else "bad"
        detail_href = _unit_detail_filename(s.get("unit"))
        rows_html += f'''<tr>
  <td class="unit"><a href="{detail_href}">{_esc(s.get("unit"))}</a></td>
  <td>{_esc(s.get("ip"))}</td>
  <td>{_esc(s.get("pipeline"))}</td>
  <td>{_esc(s.get("action"))}</td>
  <td>{_esc(s.get("cycle"))}</td>
  <td>{_esc(s.get("timestamp"))}</td>
  <td class="{'ok' if reachable_ok else 'bad'}">{"Yes" if reachable_ok else "No"}</td>
  <td class="{'ok' if s.get('seek_enumeration') == 'OK' else 'bad'}">{_esc(s.get("seek_enumeration"))}</td>
  <td class="{img_class}">{_esc(s.get("image_desc"))}</td>
  <td>{_esc(s.get("pipeline_restarts"))}</td>
  <td>{_esc(s.get("self_recovered"))}</td>
  <td>{_esc(s.get("wrapper_alive"))}</td>
  <td>{_esc(s.get("cpu_pct"))}%</td>
  <td>{_esc(s.get("cpu_temp"))}&deg;C</td>
  <td>{_esc(s.get("throttled_meaning"))}</td>
  <td class="{status_class}">{_esc(status_text)}</td>
</tr>'''

    html = f'''<!doctype html>
<html>
<head>
<meta charset="utf-8">
<title>Pipeline Test - Live Status</title>
<style>{_DASHBOARD_CSS}</style>
</head>
<body>
<h1>Pipeline Test - Live Status</h1>
<div class="updated">Rewritten fresh every cycle by whichever unit finishes next - reload this page (F5) to see the latest. Click a unit for its individual trend charts. Generated {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.</div>
{'<p class="empty">No status snapshots yet - waiting for the first cycle to complete.</p>' if not snapshots else f"""<table>
<tr><th>Unit</th><th>IP</th><th>Pipeline</th><th>Action</th><th>Cycle</th><th>Last Update</th>
<th>Reachable</th><th>Seek Enum</th><th>image.pgm</th><th>Restarts (24h)</th><th>Self-Recovered</th>
<th>Wrapper Alive</th><th>CPU %</th><th>CPU Temp</th><th>Throttled</th><th>Status</th></tr>
{rows_html}
</table>"""}
</body>
</html>
'''
    DASHBOARD_PATH.write_text(html, encoding="utf-8")


def run_cycle(run_id, cycle_num, ip, user, password, unit_label, action, pipeline, log_path):
    cycle_start = time.monotonic()
    print(f"\n=== Cycle {cycle_num} | {unit_label} | {action} | {datetime.now().strftime('%H:%M:%S')} ===")

    if action == "reboot":
        ok, err = trigger_reboot(ip, user, password)
        if not ok:
            print(f"  reboot trigger failed: {err} (logging and continuing anyway)")
        print(f"  sleeping to T+{REBOOT_CHECK_INTERVAL_SEC}s...")
        time.sleep(REBOOT_CHECK_INTERVAL_SEC)
        checkpoint_desc = f"T+{REBOOT_CHECK_INTERVAL_SEC}s"
    elif action == "restart":
        restart_label = "pipeline_restart.sh" if pipeline == "old" else "systemctl restart seekcamera-gstreamer"
        ok, restart_out = trigger_restart_blocking(ip, user, password, pipeline)
        if not ok:
            print(f"  {restart_label} failed/timed out: {restart_out} (logging and continuing anyway)")
        else:
            print(f"  {restart_label} completed")
        checkpoint_desc = f"immediately after {restart_label} returned"
    else:  # none - Test 0, steady-state, no action taken at all
        checkpoint_desc = "passive check, no reboot/restart triggered"

    reachable = False
    reason = ""
    fields = {}
    try:
        fields = run_check(ip, user, password)
        reachable = True
        print(f"  reachable - enumeration={fields.get('SEEK_ENUMERATION', '?')}, "
              f"image={fields.get('IMAGE_PGM', '?')}")
    except Exception as e:
        reason = f"{type(e).__name__}: {e}"
        print(f"  NOT reachable at {checkpoint_desc}: {reason}")

    row = [
        run_id,
        cycle_num,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        unit_label,
        action,
        "Yes" if reachable else "No",
        reason,
        fields.get("SEEK_ENUMERATION", ""),
        fields.get("ENUM_FAIL_TIME_SEC", ""),
        fields.get("USB_ERRORS", ""),
        fields.get("USB_RESETS", ""),
        fields.get("ETH_FLAP_COUNT", ""),
        fields.get("IMAGE_SIZE", ""),
        fields.get("IMAGE_AGE_SEC", ""),
        "Yes" if fields.get("IMAGE_OK") == "1" else "No",
        fields.get("PIPELINE_TYPE", ""),
        fields.get("PIPELINE_VERSION", ""),
        fields.get("PIPELINE_RESTARTS", ""),
        fields.get("SELF_RECOVERED_ERRORS", ""),
        fields.get("WRAPPER_ALIVE", ""),
        _uptime_hours(fields.get("UPTIME_SEC", "")),
        fields.get("TOTAL_CPU_PCT", ""),
        fields.get("PIPELINE_CPU_PCT", ""),
        fields.get("CALIBRATION", ""),
        fields.get("TEMP_C", ""),
        fields.get("THROTTLED", ""),
        decode_throttled(fields.get("THROTTLED", "")),
        fields.get("PORT_IMGSERV", ""),
        fields.get("PORT_RTSP_8554", ""),
        fields.get("NETWORK", ""),
        fields.get("MEM_LINE", ""),
        fields.get("OVERLAY_ACTIVE", ""),
        fields.get("OVERLAY_USAGE", ""),
        fields.get("STATUS", ""),
        fields.get("DMESG_TAIL", ""),
    ]
    log_row(log_path, row, reachable)
    snapshot = write_status_snapshot(unit_label, ip, pipeline, action, run_id, cycle_num, reachable, fields, log_path)
    render_unit_detail(snapshot)
    render_dashboard()

    # pad to the nominal total cycle length before the next action fires - uses
    # the real elapsed time for this cycle (not a guess), so cadence stays
    # accurate over long runs instead of drifting
    target = {"reboot": REBOOT_CYCLE_LENGTH_SEC, "restart": RESTART_CYCLE_LENGTH_SEC,
               "none": NONE_CYCLE_LENGTH_SEC}[action]
    elapsed = time.monotonic() - cycle_start
    remaining = target - elapsed
    if remaining > 0:
        time.sleep(remaining)


def _sanitize_for_path(s: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]", "_", s)


def _esc(s):
    return str(s if s is not None else "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _fmt_dt(t):
    """Display-only date format (month/day/year) - the master log itself
    keeps storing/parsing timestamps as %Y-%m-%d for sortability and so
    datetime.strptime in generate_run_report keeps working on old rows;
    this is purely for what a reader sees in the report."""
    return t.strftime("%m/%d/%Y %H:%M:%S")


def _usb_dmesg_excerpt(dmesg_tail, n=25):
    """Pulls just the USB/enumeration-relevant lines out of a captured dmesg
    tail (which is mostly unrelated boot noise) - lets a Seek enumeration
    failure's Findings section show the actual kernel evidence (descriptor
    read timeouts, port power-cycle attempts, address-not-accepted, etc.)
    instead of just "it failed." Keeps the last n matching lines, closest to
    the actual failure moment."""
    lines = (dmesg_tail or "").split("|")
    keep = [l.strip() for l in lines if re.search(r"\busb\b|xhci|enumerate", l, re.I)]
    return keep[-n:]


def _x_ticks(t_min, t_span_sec, pad_l, plot_w, plot_h, y_top, n=5):
    """Shared evenly-spaced wall-clock tick marks/labels for the X axis -
    used by both chart types so a reader can line up a reachability drop
    with the same moment on a metric chart below it. First/last ticks are
    explicitly labeled "test start"/"test end" (in addition to the actual
    time) so it's unambiguous these are the run's boundaries, not just
    another sample point."""
    out = ""
    for i in range(n):
        frac = i / (n - 1)
        t = t_min + timedelta(seconds=t_span_sec * frac)
        x = pad_l + frac * plot_w
        anchor = "start" if i == 0 else ("end" if i == n - 1 else "middle")
        label = t.strftime("%H:%M:%S")
        if i == 0:
            label = f"test start {label}"
        elif i == n - 1:
            label = f"test end {label}"
        out += f'<line x1="{x:.1f}" y1="{y_top}" x2="{x:.1f}" y2="{y_top + plot_h}" stroke="#f0f0f0" stroke-width="1"/>\n'
        out += f'<text x="{x:.1f}" y="{y_top + plot_h + 16}" font-size="10" text-anchor="{anchor}" fill="#666">{label}</text>\n'
    return out


def _svg_line_chart(title, timestamps, values, color, width=760, height=170, zero_floor=False):
    """Minimal dependency-free SVG line chart (wall-clock time on X). No
    JS/CDN - has to open standalone on a Windows machine with no guaranteed
    internet. Missing values (unreachable cycles) break the line into
    separate polyline segments rather than being plotted as 0, which would
    misrepresent a gap in data as a real reading. X axis is actual elapsed
    time, not cycle index, since cycle length can vary slightly cycle-to-
    cycle (padding drift) and a reader wants to know "when," not "which
    numbered attempt."

    zero_floor=True pins the y-axis bottom to 0 instead of auto-scaling to
    the data's own min - for a mostly-zero event-count metric (USB errors/
    resets, eth0 flaps), auto-scaling produces a nonsensical axis (e.g. all
    values 0 -> padded to -1..1, a metric that can never go negative) and
    makes an occasional spike look like gradual wandering instead of the
    blip-from-baseline it actually is. Continuous gauge metrics (CPU temp,
    CPU%) keep auto-scaling - a 0-70C axis would flatten the real variation
    in a temperature that only ever ranges ~50-70C into a useless sliver."""
    pad_l, pad_r, pad_t, pad_b = 50, 20, 26, 34
    plot_w = width - pad_l - pad_r
    plot_h = height - pad_t - pad_b

    known = [v for v in values if v is not None]
    if not known or not timestamps:
        return f'<div class="chart"><h3>{title}</h3><p class="empty">No data</p></div>'

    t_min, t_max = min(timestamps), max(timestamps)
    if zero_floor:
        y_min, y_max = 0, max(known)
        if y_max <= 0:
            y_max = 1
    else:
        y_min, y_max = min(known), max(known)
        if y_min == y_max:
            y_min -= 1
            y_max += 1
    t_span = (t_max - t_min).total_seconds() or 1
    y_span = (y_max - y_min) or 1

    def sx(t):
        return pad_l + (t - t_min).total_seconds() / t_span * plot_w

    def sy(v):
        return pad_t + plot_h - (v - y_min) / y_span * plot_h

    segments, current = [], []
    for t, v in zip(timestamps, values):
        if v is None:
            if current:
                segments.append(current)
                current = []
            continue
        current.append((t, v))
    if current:
        segments.append(current)

    polylines = "\n".join(
        '<polyline points="{}" fill="none" stroke="{}" stroke-width="2"/>'.format(
            " ".join(f"{sx(t):.1f},{sy(v):.1f}" for t, v in seg), color
        )
        for seg in segments
    )

    # Invisible per-point hit-circles carrying a native SVG <title> (browser
    # tooltip on hover, no JS/CDN needed - keeps this dependency-free per the
    # module docstring). No visible dot - just the line - the hit-circle
    # only needs to be there for hovering, not drawn.
    points = "\n".join(
        f'<circle cx="{sx(t):.1f}" cy="{sy(v):.1f}" r="6" fill="transparent">'
        f'<title>{_fmt_dt(t)}: {v:g}</title></circle>'
        for seg in segments for t, v in seg
    )

    gridlines = ""
    for frac in (0, 0.5, 1):
        y = pad_t + plot_h * (1 - frac)
        val = y_min + y_span * frac
        gridlines += (
            f'<line x1="{pad_l}" y1="{y:.1f}" x2="{width - pad_r}" y2="{y:.1f}" '
            f'stroke="#e0e0e0" stroke-width="1"/>\n'
            f'<text x="{pad_l - 6}" y="{y + 4:.1f}" font-size="10" text-anchor="end" fill="#666">{val:.1f}</text>\n'
        )

    xticks = _x_ticks(t_min, t_span, pad_l, plot_w, plot_h, pad_t)

    return f'''<div class="chart">
  <h3>{title}</h3>
  <svg viewBox="0 0 {width} {height}" width="100%" height="{height}">
    {gridlines}
    {xticks}
    {polylines}
    {points}
  </svg>
</div>'''


_STATE_COLORS = {"ok": "#4caf50", "fail": "#e53935", "blip": "#bdbdbd"}


def _reachability_strip(timestamps, states, title="Reachable at checkpoint", width=760, height=44):
    """states: one of "ok"/"fail"/"blip" per cycle. "blip" (gray) is for a
    cycle with no real data - e.g. the checkpoint SSH itself timed out, so
    camera-check-reboot-test.sh never ran and every field came back blank.
    That's a missed checkpoint, not evidence the thing being charted actually
    failed - conflating the two made a single SSH timeout look identical to a
    real Seek enumeration failure (see PT1 cycle 221, 2026-08-21 - blank
    Seek Camera Enumeration counted as FAILED when the row was simply
    unreachable that cycle)."""
    n = max(len(timestamps), 1)
    seg_w = width / n
    rects = "\n".join(
        f'<rect x="{i * seg_w:.1f}" y="0" width="{seg_w + 0.5:.1f}" height="24" '
        f'fill="{_STATE_COLORS[s]}"><title>{_fmt_dt(timestamps[i])}: {s}</title></rect>'
        for i, s in enumerate(states)
    )
    t_min, t_max = min(timestamps), max(timestamps)
    t_span = (t_max - t_min).total_seconds() or 1
    xticks = _x_ticks(t_min, t_span, 0, width, 0, 24, n=5)
    return f'''<div class="chart">
  <h3>{title}</h3>
  <svg viewBox="0 0 {width} {height}" width="100%" height="{height}">{rects}{xticks}</svg>
</div>'''


def generate_run_report(log_path, run_id, unit_label, action):
    """Writes a per-run report folder: a spreadsheet filtered to just this
    run's rows (Cycle numbers reset every run, so the shared master log
    needs Run ID to tell runs apart - this pulls one run back out into its
    own file) plus a dependency-free HTML trend chart. Reads back from the
    saved log rather than tracking rows in memory, so this produces the
    same result whether the run finished normally, hit --cycles/
    --duration-min, or was stopped early with Ctrl+C."""
    wb = openpyxl.load_workbook(log_path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    run_id_idx = header.index("Run ID")
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[run_id_idx] == run_id]
    if not rows:
        print("No rows found for this run - skipping report generation.")
        return

    # run_id is already MMDDYY_HHMMSS (see main()) - directly usable as a
    # folder name segment, no reformatting needed.
    run_tag = _sanitize_for_path(f"{unit_label}_{run_id}")
    # Anchored on this script's own directory, not log_path.parent - the
    # master log can live anywhere (e.g. logs/RebootLoopTest_PT1.xlsx), but
    # reports/ should stay one stable, predictable top-level location
    # regardless of where --log points.
    report_dir = Path(__file__).parent / "reports" / run_tag
    report_dir.mkdir(parents=True, exist_ok=True)

    # -- filtered spreadsheet, styled the same as the master log --
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Reboot Loop Test"
    out_ws.append(COLUMNS)
    for cell in out_ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    reachable_col = header.index("Reachable at checkpoint")
    for r in rows:
        out_ws.append(list(r))
        fill = GREEN_FILL if r[reachable_col] == "Yes" else RED_FILL
        row_idx = out_ws.max_row
        for col_idx in range(1, len(COLUMNS) + 1):
            cell = out_ws.cell(row=row_idx, column=col_idx)
            cell.font = BASE_FONT
            cell.border = BORDER
            cell.fill = fill
    out_ws.freeze_panes = "A2"
    xlsx_path = report_dir / f"{run_tag}.xlsx"
    out_wb.save(xlsx_path)

    # -- HTML trend report --
    timestamp_idx = header.index("Timestamp")
    timestamps = [datetime.strptime(r[timestamp_idx], "%Y-%m-%d %H:%M:%S") for r in rows]
    reachable_flags = [r[reachable_col] == "Yes" for r in rows]

    def col(col_name):
        return [r[header.index(col_name)] for r in rows]

    def numeric_series(col_name):
        idx = header.index(col_name)
        out = []
        for r in rows:
            try:
                out.append(float(r[idx]))
            except (TypeError, ValueError):
                out.append(None)
        return out

    def numeric_stats(col_name):
        vals = [v for v in numeric_series(col_name) if v is not None]
        if not vals:
            return None
        return {"min": min(vals), "avg": sum(vals) / len(vals), "max": max(vals), "sum": sum(vals)}

    def make_states(raw_values, reachable, is_ok):
        """"blip" (gray, not a real fail) when there's no real data for this
        field - either the checkpoint SSH itself was unreachable (see PT1
        cycle 221, 2026-08-21 - a missed checkpoint originally got counted
        as a Seek enumeration failure it never had), OR - rarer, but real -
        the SSH connected fine (Reachable=Yes) but the check script's own
        output came back blank for this cycle anyway (see PT2 cycle 429,
        2026-08-21: Reachable=Yes yet Seek/Pipeline Type/Status were all
        None - a partial/incomplete read, not a real result). Checking
        Reachable alone missed that second case and let it masquerade as
        this run's "first" Seek enumeration failure, when the real first
        failure was actually the next real reading two cycles later."""
        states = []
        for v, r in zip(raw_values, reachable):
            if not r or v in (None, ""):
                states.append("blip")
            else:
                states.append("ok" if is_ok(v) else "fail")
        return states

    def first_and_still(states):
        """First index (ignoring blips) where state is "fail", and whether
        every non-blip state from there to the end of the run is also
        "fail" - lets the summary distinguish a one-off blip from a failure
        that never recovered for the rest of the run (the pattern that
        actually matters - see PT1NEW/PT2 2026-08-20/21 overnight run,
        where Seek enumeration failed once and never came back for
        7-8+ hours)."""
        non_blip = [(i, s) for i, s in enumerate(states) if s != "blip"]
        fail_idxs = [i for i, s in non_blip if s == "fail"]
        if not fail_idxs:
            return None, False
        first_idx = fail_idxs[0]
        tail = [s for i, s in non_blip if i >= first_idx]
        return first_idx, all(s == "fail" for s in tail)

    def status_row(label, states, fail_word="FAILED"):
        n_fail = states.count("fail")
        n_blip = states.count("blip")
        n_checked = len(states) - n_blip
        blip_note = f" ({n_blip} cycle(s) had no data - checkpoint unreachable)" if n_blip else ""
        if n_fail == 0:
            return f"<tr><th>{label}</th><td>Never {fail_word.lower()} - {n_checked}/{n_checked} checked OK{blip_note}</td></tr>"
        first_idx, stuck = first_and_still(states)
        detail = f"{fail_word} {n_fail}/{n_checked} checked cycles, first at {_fmt_dt(timestamps[first_idx])}"
        if stuck:
            detail += " - never recovered for the rest of the run"
        detail += blip_note
        return f"<tr><th>{label}</th><td>{detail}</td></tr>"

    seek_col = col("Seek Camera Enumeration")
    seek_states = make_states(seek_col, reachable_flags, lambda v: v == "OK")
    image_col = col("image.pgm Healthy")
    image_states = make_states(image_col, reachable_flags, lambda v: v == "Yes")

    # -- Findings: real dmesg evidence from the first Seek enumeration
    # failure, not just "it failed" - see PT1NEW 2026-08-21 cycle 195 for
    # why this matters (descriptor read timeouts -> kernel's own port
    # power-cycle attempt -> still failing -> unable to enumerate; this is
    # a kernel/USB-controller-level event, below either pipeline's reach).
    findings_html = ""
    seek_first_idx, seek_stuck = first_and_still(seek_states)
    if seek_first_idx is not None:
        dmesg_idx = header.index("Raw dmesg tail")
        usb_lines = _usb_dmesg_excerpt(rows[seek_first_idx][dmesg_idx])
        if usb_lines:
            ts_str = _fmt_dt(timestamps[seek_first_idx])
            findings_html = f'''
<h2>Findings: Seek Enumeration failure (first at {ts_str})</h2>
<p>USB/enumeration-relevant kernel log lines captured at that checkpoint - the actual
evidence for what the kernel saw, not just the fact that enumeration failed:</p>
<pre>{chr(10).join(_esc(l) for l in usb_lines)}</pre>
'''

    charts_html = _reachability_strip(timestamps, ["ok" if r else "fail" for r in reachable_flags], "Reachable at checkpoint")
    charts_html += "\n" + _reachability_strip(timestamps, seek_states, "Seek Camera Enumeration (green=OK, red=FAILED, gray=no data)")
    charts_html += "\n" + _reachability_strip(timestamps, image_states, "image.pgm Healthy (green=Yes, red=No, gray=no data)")

    metrics = [
        ("CPU Temp (C)", "#e65100"),
        ("Total CPU %", "#00838f"),
        ("Pipeline CPU %", "#558b2f"),
    ]
    count_metrics = [
        ("USB Communication Errors", "#1565c0"),
        ("USB Resets", "#6a1b9a"),
        ("Ethernet Flap Count", "#2e7d32"),
    ]
    for col_name, color in metrics:
        charts_html += "\n" + _svg_line_chart(col_name, timestamps, numeric_series(col_name), color)
    for col_name, color in count_metrics:
        charts_html += "\n" + _svg_line_chart(col_name, timestamps, numeric_series(col_name), color, zero_floor=True)

    total = len(rows)
    reachable_count = sum(reachable_flags)

    ptype_col = col("Pipeline Type")
    pipeline_types = Counter(v for v in ptype_col if v)
    ptype_flag = ""
    ptype_flag_class = "note"
    old_idxs = []  # populated below only when there's a real mix to analyze
    if len(pipeline_types) <= 1:
        ptype_str = next(iter(pipeline_types), "n/a")
    else:
        # A single isolated "old" reading (correct type right before AND after)
        # is just the check landing in the ~1-2s window right after `systemctl
        # restart` fires, before systemd reports the unit active again - the
        # pipeline never actually switched, the check just misread it that one
        # cycle (confirmed on PT2NEW 2026-08-21: Pipeline Restarts kept
        # climbing normally through every one of these, no freeze).
        #
        # A RUN of several consecutive "old" readings means something else,
        # and it's action-dependent: for a restart-action test, that's the
        # real StartLimitBurst symptom (systemctl restart being rate-limited -
        # see WEEKEND_TEST_FINDINGS.md). For a reboot-action test there is no
        # systemctl restart happening at all - StartLimitBurst doesn't apply.
        # There, a run of "old" readings almost always means the pipeline
        # service just took longer than the T+1:30 checkpoint to reach
        # "active" on that particular boot - and if those cycles overlap with
        # an ongoing Seek Enumeration failure (confirmed on PT1NEW
        # 2026-08-21: all 19 "old" readings were also all 19 image.pgm
        # Unhealthy readings, and every one fell inside the existing Seek
        # enumeration outage), that's not a separate incident - it's the
        # service repeatedly retrying to find a camera that's already known
        # to be missing.
        runs, cur = [], None
        for v in ptype_col:
            if cur and cur[0] == v:
                cur[1] += 1
            else:
                cur = [v, 1]
                runs.append(cur)
        old_run_lengths = [n for v, n in runs if v == "old"]
        max_old_run = max(old_run_lengths, default=0)
        old_idxs = [i for i, v in enumerate(ptype_col) if v == "old"]
        overlap_with_seek_fail = sum(1 for i in old_idxs if seek_states[i] == "fail")

        if max_old_run >= 3 and action == "restart":
            ptype_str = ", ".join(f"{k}: {v}" for k, v in pipeline_types.most_common())
            ptype_flag_class = "warn"
            ptype_flag = (f" &#9888; sustained old-pipeline reading ({max_old_run} consecutive cycles) - "
                           f"StartLimitBurst-style stall, see WEEKEND_TEST_FINDINGS.md")
        elif len(old_idxs) and overlap_with_seek_fail == len(old_idxs):
            # Fully explained by the Seek Enumeration failure already shown
            # below (and in Findings) - no separate note needed here.
            ptype_str = pipeline_types.most_common(1)[0][0]
        elif max_old_run >= 3:
            ptype_str = ", ".join(f"{k}: {v}" for k, v in pipeline_types.most_common())
            ptype_flag_class = "warn"
            ptype_flag = (f" &#9888; sustained old-pipeline reading ({max_old_run} consecutive cycles) - "
                           f"pipeline took longer than the checkpoint to come up on {len(old_run_lengths)} boot(s), "
                           f"not linked to a Seek enumeration failure - worth a closer look")
        else:
            ptype_str = pipeline_types.most_common(1)[0][0]  # dominant/real type - "old" here was never real
            ptype_flag = (f" ({len(old_run_lengths)} isolated single-cycle misread(s), never actually switched "
                           f"pipelines - checkpoint landed mid-restart; restarts count kept climbing normally "
                           f"around each one)")

    # Exclude blip rows (no real data - checkpoint unreachable, or reachable
    # but the check's own output came back blank anyway, see make_states
    # above) - those aren't a real calibration state, same class of bug
    # fixed for Seek Enumeration/image.pgm above. Also exclude cycles where
    # Pipeline Type itself was a misread (not a real old-pipeline reading,
    # see the ptype block above) - "n/a (old pipeline)" on those rows is
    # just a side effect of the same misread, not a real calibration state.
    ptype_misread_idxs = set(old_idxs) if ptype_flag_class != "warn" else set()
    calib_counts = Counter()
    calib_blips = 0
    for i, (v, reachable) in enumerate(zip(col("Calibration"), reachable_flags)):
        if i in ptype_misread_idxs:
            continue
        if not reachable or v in (None, ""):
            calib_blips += 1
            continue
        calib_counts[v.split(" - ")[0]] += 1
    calib_blip_note = f" ({calib_blips} cycle(s) had no data - checkpoint unreachable)" if calib_blips else ""
    if len(calib_counts) == 1:
        calib_str = next(iter(calib_counts)) + calib_blip_note
    else:
        calib_str = (", ".join(f"{k}: {v}" for k, v in calib_counts.most_common()) or "n/a") + calib_blip_note

    def rolling_24h_str(col_name):
        """Both of these columns are rolling 24h counters (re-read from
        journalctl --since "-24 hours" fresh every cycle), so the raw
        first/last values themselves aren't the useful number - the delta
        between them is (~how many happened during this run). That delta is
        approximate, not exact (a restart from before the run can also age
        out of the 24h window while the run is in progress, undercounting
        slightly) - full detail in the hover tooltip instead of cluttering
        the main line."""
        vals = [v for v in numeric_series(col_name) if v is not None]
        if not vals:
            return "n/a"
        delta = vals[-1] - vals[0]
        detail = f"rolling 24h count: {vals[0]:.0f} at start, {vals[-1]:.0f} at end - delta is approximate"
        return f'<span title="{detail}">~{delta:.0f} during this run</span>'

    restarts_str = rolling_24h_str("Pipeline Restarts (24h)")
    self_rec_str = rolling_24h_str("Self-Recovered Errors (24h)")

    # Status text embeds live counters (e.g. "USB UNSTABLE - 334 resets..."),
    # which climb almost every cycle - grouping by the exact string means a
    # single ongoing issue fragments into dozens of near-duplicate 1-cycle
    # "statuses" instead of one real bucket. Normalize digits out for
    # grouping/counting, but keep one real example string per bucket to
    # display so it's still readable. Blip rows (blank Status - no real data
    # for that cycle, same as everywhere else in this report) are excluded
    # from the breakdown entirely rather than showing up as a confusing
    # count-only row with empty text - see PT1 cycle 221, 2026-08-21.
    def normalize_status(s):
        return re.sub(r"\d+", "N", s or "")

    status_examples = {}
    status_blips = 0
    for s in col("Status"):
        if not s:
            status_blips += 1
            continue
        status_examples.setdefault(normalize_status(s), s)
    status_counts = Counter(normalize_status(s) for s in col("Status") if s)
    status_rows_html = "\n".join(
        f'<tr><td>{cnt}</td><td>{_esc(status_examples[key][:140])}</td></tr>'
        for key, cnt in status_counts.most_common(8)
    )
    if status_blips:
        status_rows_html += (f'<tr><td>{status_blips}</td><td class="note">'
                              f'(no data - checkpoint unreachable or incomplete)</td></tr>')

    def metric_row(label, col_name, unit=""):
        """No sum column - it was meaningless for two different reasons
        depending on the metric: instantaneous readings like temperature/
        CPU% can't be meaningfully summed at all, and for restart-action
        tests (dmesg persists across the whole run, doesn't reset per
        cycle) USB error/reset/flap counts are themselves running totals
        already, so summing them again just multiplies the same events by
        however many cycles they stayed visible in the tail window - not a
        real count of anything. Min/Avg/Max stay meaningful either way."""
        s = numeric_stats(col_name)
        if s is None:
            return f"<tr><td>{label}</td><td colspan=3>n/a</td></tr>"
        return (f"<tr><td>{label}</td><td>{s['min']:.1f}{unit}</td><td>{s['avg']:.1f}{unit}</td>"
                f"<td>{s['max']:.1f}{unit}</td></tr>")

    html = f'''<!doctype html>
<html>
<head>
<meta charset="utf-8">
<title>{unit_label} - {run_id} - Reboot/Restart Loop Test Report</title>
<style>
  body {{ font-family: Arial, sans-serif; margin: 24px; color: #222; }}
  h1 {{ font-size: 20px; }}
  h2 {{ font-size: 15px; margin-top: 28px; border-bottom: 1px solid #ddd; padding-bottom: 4px; }}
  table {{ border-collapse: collapse; margin-bottom: 12px; }}
  td, th {{ padding: 4px 12px; text-align: left; border-bottom: 1px solid #eee; font-size: 13px; }}
  th {{ white-space: nowrap; }}
  .chart {{ margin-bottom: 28px; }}
  .chart h3 {{ margin: 0 0 6px 0; font-size: 14px; color: #444; }}
  .empty {{ color: #999; font-size: 12px; }}
  .warn {{ color: #c62828; font-weight: bold; }}
  .note {{ color: #777; }}
  pre {{ background: #f7f7f7; border: 1px solid #ddd; border-radius: 4px; padding: 10px 12px;
         font-size: 11px; line-height: 1.5; overflow-x: auto; }}
</style>
</head>
<body>
<h1>{unit_label} - {action} loop test</h1>

<h2>Overview</h2>
<table>
<tr><th>Run ID</th><td>{run_id}</td></tr>
<tr><th>Cycles logged</th><td>{total}</td></tr>
<tr><th>Time span</th><td>{_fmt_dt(timestamps[0])} - {_fmt_dt(timestamps[-1])}</td></tr>
<tr><th>Reachable at checkpoint</th><td>{reachable_count} / {total} ({100 * reachable_count / total:.1f}%)</td></tr>
{status_row("Seek Camera Enumeration", seek_states)}
{status_row("image.pgm Healthy", image_states, fail_word="Unhealthy")}
<tr><th>Pipeline Type</th><td>{ptype_str}<span class="{ptype_flag_class}">{ptype_flag}</span></td></tr>
<tr><th>Pipeline Restarts (24h)</th><td>{restarts_str}</td></tr>
<tr><th>Self-Recovered Errors (24h)</th><td>{self_rec_str}</td></tr>
<tr><th>Calibration</th><td>{calib_str}</td></tr>
<tr><th>Spreadsheet</th><td>{xlsx_path.name}</td></tr>
</table>
{findings_html}
<h2>Numeric metrics</h2>
<table>
<tr><th>Metric</th><th>Min</th><th>Avg</th><th>Max</th></tr>
{metric_row("CPU Temp (C)", "CPU Temp (C)", "&deg;C")}
{metric_row("Total CPU %", "Total CPU %", "%")}
{metric_row("Pipeline CPU %", "Pipeline CPU %", "%")}
{metric_row("USB Communication Errors", "USB Communication Errors")}
{metric_row("USB Resets", "USB Resets")}
{metric_row("Ethernet Flap Count", "Ethernet Flap Count")}
</table>

<h2>Status breakdown (top 8)</h2>
<table>
<tr><th>Count</th><th>Status</th></tr>
{status_rows_html}
</table>

<h2>Charts</h2>
{charts_html}
</body>
</html>
'''
    html_path = report_dir / "trend_report.html"
    html_path.write_text(html, encoding="utf-8")

    print(f"\nRun report written to {report_dir}")
    print(f"  spreadsheet: {xlsx_path.name}")
    print(f"  trend chart: {html_path.name}")


def main():
    parser = argparse.ArgumentParser(description="Reboot/restart loop test - PT vs PT RamOS")
    parser.add_argument("--ip", required=True)
    parser.add_argument("--user", default="camera")
    parser.add_argument("--password", required=True)
    parser.add_argument("--unit", required=True, help='e.g. "PT" or "PT RamOS"')
    parser.add_argument("--action", choices=["reboot", "restart", "none"], required=True,
                         help="reboot = Test 1 (2-min cycle), restart = Test 2, pipeline_restart.sh (60s cycle), "
                              "none = Test 0 steady-state (5-min cycle, passive check only, no action triggered)")
    parser.add_argument("--pipeline", choices=["old", "new"], required=True,
                         help="which pipeline this device is running - determines the restart "
                              "command for --action restart (old: pipeline_restart.sh, "
                              "new: systemctl restart seekcamera-gstreamer); ignored for --action reboot "
                              "but still required for consistency/logging clarity")
    parser.add_argument("--cycles", type=int, default=None, help="stop after N cycles")
    parser.add_argument("--duration-min", type=float, default=None, help="stop after ~N minutes")
    parser.add_argument("--log", default=None, help="output xlsx path (default: RebootLoopTest.xlsx next to this script)")
    args = parser.parse_args()

    log_path = Path(args.log) if args.log else Path(__file__).parent / "RebootLoopTest.xlsx"

    # One ID per script invocation, not per cycle - Cycle numbers restart at 1
    # every run, so appending a second run to the same log file would otherwise
    # produce duplicate Cycle values with no way to tell the runs apart without
    # cross-referencing Timestamp by hand. MMDDYY_HHMMSS (not ISO) so it can be
    # used directly as a report folder name segment with no reformatting/
    # duplicate-date handling needed - see generate_run_report().
    run_id = datetime.now().strftime("%m%d%y_%H%M%S")
    print(f"Run ID: {run_id}")

    start_time = time.monotonic()
    cycle = 0
    try:
        while True:
            cycle += 1
            run_cycle(run_id, cycle, args.ip, args.user, args.password, args.unit, args.action, args.pipeline, log_path)

            if args.cycles and cycle >= args.cycles:
                print(f"\nReached {args.cycles} cycles, stopping.")
                break
            if args.duration_min and (time.monotonic() - start_time) >= args.duration_min * 60:
                print(f"\nReached ~{args.duration_min} min, stopping.")
                break
    except KeyboardInterrupt:
        print(f"\nStopped by user after {cycle} cycles.")

    print(f"Results logged to {log_path}")
    generate_run_report(log_path, run_id, args.unit, args.action)


if __name__ == "__main__":
    main()
